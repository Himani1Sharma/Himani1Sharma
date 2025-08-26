from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timedelta
import io
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import matplotlib
matplotlib.use("Agg")  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import base64

from src.services.weather_service import WeatherService
from src.models.weather import WeatherData

weather_bp = Blueprint("weather", __name__)

@weather_bp.route("/health", methods=["GET"])
def health_check():
    """
    Health Check
    This endpoint provides the current status of the Weather Backend Service.
    ---
    responses:
      200:
        description: Service is healthy.
        schema:
          type: object
          properties:
            status:
              type: string
              example: healthy
            service:
              type: string
              example: Weather Backend Service
            timestamp:
              type: string
              format: date-time
            database:
              type: object
              properties:
                status:
                  type: string
                  example: connected
                total_records:
                  type: integer
      500:
        description: Service is unhealthy or an error occurred.
        schema:
          type: object
          properties:
            status:
              type: string
              example: unhealthy
            service:
              type: string
              example: Weather Backend Service
            timestamp:
              type: string
              format: date-time
            error:
              type: string
    tags:
      - System
    """
    try:
        # Check database connectivity
        record_count = WeatherData.query.count()

        return jsonify({
            "status": "healthy",
            "service": "Weather Backend Service",
            "timestamp": datetime.utcnow().isoformat(),
            "database": {
                "status": "connected",
                "total_records": record_count
            }
        }), 200

    except Exception as e:
        return jsonify({
            "status": "unhealthy",
            "service": "Weather Backend Service",
            "timestamp": datetime.utcnow().isoformat(),
            "error": str(e)
        }), 500

@weather_bp.route("/weather-report", methods=["GET"])
def get_weather_report():
    """
    Fetch and Store Weather Data
    This endpoint fetches time-series temperature and humidity data for the past 2 days
    from the Open-Meteo Archive API and stores it in the SQLite database.
    ---
    parameters:
      - name: lat
        in: query
        type: number
        format: float
        required: true
        description: Latitude coordinate (e.g., 47.37). Must be between -90 and 90.
      - name: lon
        in: query
        type: number
        format: float
        required: true
        description: Longitude coordinate (e.g., 8.55). Must be between -180 and 180.
    responses:
      200:
        description: Successfully fetched and stored weather data.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            message:
              type: string
              example: Successfully fetched and stored 48 weather records
            location:
              type: object
              properties:
                latitude:
                  type: number
                longitude:
                  type: number
            records_count:
              type: integer
            data_range:
              type: object
              properties:
                start:
                  type: string
                  format: date-time
                end:
                  type: string
                  format: date-time
            sample_data:
              type: array
              items:
                type: object
      400:
        description: Missing or invalid parameters.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      500:
        description: Internal server error or API fetch failed.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
    tags:
      - Weather Data
    """
    try:
        # Get query parameters
        lat = request.args.get("lat", type=float)
        lon = request.args.get("lon", type=float)

        # Validate parameters
        if lat is None or lon is None:
            return jsonify({
                "error": "Missing required parameters",
                "message": "Both lat and lon parameters are required"
            }), 400

        # Validate coordinate ranges
        if not (-90 <= lat <= 90):
            return jsonify({
                "error": "Invalid latitude",
                "message": "Latitude must be between -90 and 90"
            }), 400

        if not (-180 <= lon <= 180):
            return jsonify({
                "error": "Invalid longitude",
                "message": "Longitude must be between -180 and 180"
            }), 400

        # Fetch data from Open-Meteo API
        try:
            api_data = WeatherService.fetch_weather_data(lat, lon, days=2)
        except Exception as e:
            return jsonify({
                "error": "API fetch failed",
                "message": str(e)
            }), 500

        # Process and store data in database
        try:
            weather_records = WeatherService.process_and_store_weather_data(api_data, lat, lon)
        except Exception as e:
            return jsonify({
                "error": "Database storage failed",
                "message": str(e)
            }), 500

        # Return success response with data summary
        return jsonify({
            "status": "success",
            "message": f"Successfully fetched and stored {len(weather_records)} weather records",
            "location": {
                "latitude": lat,
                "longitude": lon
            },
            "records_count": len(weather_records),
            "data_range": {
                "start": weather_records[0].timestamp.isoformat() if weather_records else None,
                "end": weather_records[-1].timestamp.isoformat() if weather_records else None
            },
            "sample_data": [record.to_dict() for record in weather_records[:5]]  # First 5 records as sample
        }), 200

    except Exception as e:
        return jsonify({
            "error": "Internal server error",
            "message": str(e)
        }), 500

@weather_bp.route("/weather-data", methods=["GET"])
def get_weather_data():
    """
    Retrieve Stored Weather Data
    This endpoint retrieves stored weather data from the database based on location and time range.
    ---
    parameters:
      - name: lat
        in: query
        type: number
        format: float
        required: false
        description: Latitude coordinate to filter by location. Must be between -90 and 90.
      - name: lon
        in: query
        type: number
        format: float
        required: false
        description: Longitude coordinate to filter by location. Must be between -180 and 180.
      - name: hours
        in: query
        type: integer
        required: false
        default: 48
        description: Number of hours to look back for data (e.g., 24, 48). Max 168 hours (1 week).
    responses:
      200:
        description: Successfully retrieved weather data.
        schema:
          type: object
          properties:
            status:
              type: string
              example: success
            records_count:
              type: integer
            hours_requested:
              type: integer
            location_filter:
              type: object
              properties:
                latitude:
                  type: number
                longitude:
                  type: number
            data:
              type: array
              items:
                type: object
                properties:
                  id:
                    type: integer
                  timestamp:
                    type: string
                    format: date-time
                  latitude:
                    type: number
                  longitude:
                    type: number
                  temperature_2m:
                    type: number
                  relative_humidity_2m:
                    type: number
                  created_at:
                    type: string
                    format: date-time
      400:
        description: Invalid hours or coordinates parameter.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      500:
        description: Internal server error.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
    tags:
      - Weather Data
    """
    try:
        # Get query parameters
        lat = request.args.get("lat", type=float)
        lon = request.args.get("lon", type=float)
        hours = request.args.get("hours", type=int, default=48)

        # Validate hours parameter
        if hours <= 0 or hours > 168:  # Max 1 week
            return jsonify({
                "error": "Invalid hours parameter",
                "message": "Hours must be between 1 and 168 (1 week)"
            }), 400

        # Fetch data based on parameters
        if lat is not None and lon is not None:
            # Validate coordinates
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                return jsonify({
                    "error": "Invalid coordinates",
                    "message": "Latitude must be between -90 and 90, longitude between -180 and 180"
                }), 400

            weather_data = WeatherService.get_weather_data_by_location(lat, lon, hours)
        else:
            weather_data = WeatherService.get_recent_weather_data(hours)

        # Convert to JSON format
        data_list = [record.to_dict() for record in weather_data]

        return jsonify({
            "status": "success",
            "records_count": len(data_list),
            "hours_requested": hours,
            "location_filter": {
                "latitude": lat,
                "longitude": lon
            } if lat is not None and lon is not None else None,
            "data": data_list
        }), 200

    except Exception as e:
        return jsonify({
            "error": "Internal server error",
            "message": str(e)
        }), 500

@weather_bp.route("/export/excel", methods=["GET"])
def export_excel():
    """
    Export Weather Data to Excel
    This endpoint generates and returns an Excel file (.xlsx) containing weather data.
    ---
    parameters:
      - name: lat
        in: query
        type: number
        format: float
        required: false
        description: Latitude coordinate to filter by location. Must be between -90 and 90.
      - name: lon
        in: query
        type: number
        format: float
        required: false
        description: Longitude coordinate to filter by location. Must be between -180 and 180.
      - name: hours
        in: query
        type: integer
        required: false
        default: 48
        description: Number of hours to look back for data (e.g., 24, 48). Max 168 hours (1 week).
    responses:
      200:
        description: Successfully generated and downloaded Excel file.
        schema:
          type: string
          format: binary
      400:
        description: Invalid hours or coordinates parameter.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      404:
        description: No data found for the specified parameters.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      500:
        description: Internal server error during Excel generation.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
    tags:
      - Export
    """
    try:
        # Get query parameters
        lat = request.args.get("lat", type=float)
        lon = request.args.get("lon", type=float)
        hours = request.args.get("hours", type=int, default=48)

        # Validate hours parameter
        if hours <= 0 or hours > 168:  # Max 1 week
            return jsonify({
                "error": "Invalid hours parameter",
                "message": "Hours must be between 1 and 168 (1 week)"
            }), 400

        # Fetch data based on parameters
        if lat is not None and lon is not None:
            # Validate coordinates
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                return jsonify({
                    "error": "Invalid coordinates",
                    "message": "Latitude must be between -90 and 90, longitude between -180 and 180"
                }), 400

            weather_data = WeatherService.get_weather_data_by_location(lat, lon, hours)
            location_info = f"Lat: {lat}, Lon: {lon}"
        else:
            weather_data = WeatherService.get_recent_weather_data(hours)
            location_info = "All locations"

        if not weather_data:
            return jsonify({
                "error": "No data found",
                "message": "No weather data available for the specified parameters"
            }), 404

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Weather Data"

        # Add title and metadata
        ws["A1"] = "Weather Data Export"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Location: {location_info}"
        ws["A3"] = f"Time Range: {hours} hours"
        ws["A4"] = f"Export Date: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
        ws["A5"] = f"Total Records: {len(weather_data)}"

        # Add headers to match required column names
        headers = ["timestamp", "temperature_2m", "relative_humidity_2m"]
        header_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_idx, record in enumerate(weather_data, header_row + 1):
            ws.cell(row=row_idx, column=1, value=record.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=2, value=record.temperature_2m)
            ws.cell(row=row_idx, column=3, value=record.relative_humidity_2m)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        # Generate filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"weather_data_{timestamp}.xlsx"

        # Return file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({
            "error": "Excel export failed",
            "message": str(e)
        }), 500
    finally:
        # Clean up temporary file
        try:
            if "temp_file" in locals():
                os.unlink(temp_file.name)
        except:
            pass

@weather_bp.route("/export/pdf", methods=["GET"])
def export_pdf():
    """
    Export Weather Data to PDF Report
    This endpoint generates and returns a PDF report with charts visualizing weather data.
    ---
    parameters:
      - name: lat
        in: query
        type: number
        format: float
        required: false
        description: Latitude coordinate to filter by location. Must be between -90 and 90.
      - name: lon
        in: query
        type: number
        format: float
        required: false
        description: Longitude coordinate to filter by location. Must be between -180 and 180.
      - name: hours
        in: query
        type: integer
        required: false
        default: 48
        description: Number of hours to look back for data (e.g., 24, 48). Max 168 hours (1 week).
    responses:
      200:
        description: Successfully generated and downloaded PDF report.
        schema:
          type: string
          format: binary
      400:
        description: Invalid hours or coordinates parameter.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      404:
        description: No data found for the specified parameters.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
      500:
        description: Internal server error during PDF generation.
        schema:
          type: object
          properties:
            error:
              type: string
            message:
              type: string
    tags:
      - Export
    """
    try:
        # Lazy import to avoid DLL issues at startup on Windows
        try:
            from weasyprint import HTML, CSS
        except Exception:
            HTML = None
            CSS = None
        # Get query parameters
        lat = request.args.get("lat", type=float)
        lon = request.args.get("lon", type=float)
        hours = request.args.get("hours", type=int, default=48)

        # Validate hours parameter
        if hours <= 0 or hours > 168:  # Max 1 week
            return jsonify({
                "error": "Invalid hours parameter",
                "message": "Hours must be between 1 and 168 (1 week)"
            }), 400

        # Fetch data based on parameters
        if lat is not None and lon is not None:
            # Validate coordinates
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                return jsonify({
                    "error": "Invalid coordinates",
                    "message": "Latitude must be between -90 and 90, longitude between -180 and 180"
                }), 400

            weather_data = WeatherService.get_weather_data_by_location(lat, lon, hours)
            location_info = f"Latitude: {lat}°, Longitude: {lon}°"
        else:
            weather_data = WeatherService.get_recent_weather_data(hours)
            location_info = "All locations"

        if not weather_data:
            return jsonify({
                "error": "No data found",
                "message": "No weather data available for the specified parameters"
            }), 404

        # Generate chart
        chart_base64 = generate_weather_chart(weather_data)

        # Generate HTML content for PDF
        html_content = generate_pdf_html(weather_data, location_info, hours, chart_base64)

        # Generate PDF
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")

        # Create CSS for styling
        css_content = """
        @page {
            size: A4;
            margin: 2cm;
        }
        body {
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
        }
        .header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 2px solid #007acc;
            padding-bottom: 20px;
        }
        .title {
            font-size: 24px;
            font-weight: bold;
            color: #007acc;
            margin-bottom: 10px;
        }
        .metadata {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .chart-container {
            text-align: center;
            margin: 30px 0;
        }
        .chart-title {
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 15px;
            color: #007acc;
        }
        .summary {
            background-color: #e9f7ff;
            padding: 15px;
            border-radius: 5px;
            margin-top: 20px;
        }
        .summary-title {
            font-size: 16px;
            font-weight: bold;
            margin-bottom: 10px;
            color: #007acc;
        }
        """

        # Try WeasyPrint first if available, else fallback to ReportLab
        try:
            if HTML is not None and CSS is not None:
                HTML(string=html_content).write_pdf(temp_file.name, stylesheets=[CSS(string=css_content)])
                temp_file.close()
            else:
                raise RuntimeError("WeasyPrint not available, using ReportLab fallback")
        except Exception:
            # Fallback: generate a simple PDF with ReportLab
            try:
                pdf_path = generate_pdf_with_reportlab(weather_data, location_info, hours, chart_base64)
                # Close the temp_file we opened for WeasyPrint, since we won't use it
                try:
                    temp_file.close()
                except:
                    pass
                temp_file = type('T', (), {'name': pdf_path})()  # shim to reuse send_file path handling
            except Exception as fallback_error:
                return jsonify({
                    "error": "PDF export failed",
                    "message": "Both WeasyPrint and ReportLab fallback failed",
                    "details": str(fallback_error)
                }), 500

        # Generate filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"weather_report_{timestamp}.pdf"

        # Return file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf"
        )

    except Exception as e:
        return jsonify({
            "error": "PDF export failed",
            "message": str(e)
        }), 500
    finally:
        # Clean up temporary file
        try:
            if "temp_file" in locals():
                os.unlink(temp_file.name)
        except:
            pass

def generate_weather_chart(weather_data):
    """
    Generate a weather chart using matplotlib and return as base64 encoded image.

    Args:
        weather_data: List of WeatherData objects

    Returns:
        Base64 encoded PNG image string
    """
    # Extract data for plotting
    timestamps = [record.timestamp for record in weather_data]
    temperatures = [record.temperature_2m for record in weather_data]
    humidities = [record.relative_humidity_2m for record in weather_data]

    # Create figure with subplots
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    fig.suptitle("Weather Data Over Time", fontsize=16, fontweight="bold")

    # Temperature plot
    ax1.plot(timestamps, temperatures, color="#ff6b6b", linewidth=2, marker="o", markersize=3)
    ax1.set_ylabel("Temperature (°C)", fontsize=12)
    ax1.grid(True, alpha=0.3)
    ax1.set_title("Temperature", fontsize=14, pad=10)

    # Add temperature statistics
    temp_min = min(temperatures)
    temp_max = max(temperatures)
    temp_avg = sum(temperatures) / len(temperatures)
    ax1.axhline(y=temp_avg, color="#ff6b6b", linestyle="--", alpha=0.7, label=f"Avg: {temp_avg:.1f}°C")
    ax1.legend()

    # Humidity plot
    ax2.plot(timestamps, humidities, color="#4ecdc4", linewidth=2, marker="s", markersize=3)
    ax2.set_ylabel("Relative Humidity (%)", fontsize=12)
    ax2.set_xlabel("Time", fontsize=12)
    ax2.grid(True, alpha=0.3)
    ax2.set_title("Relative Humidity", fontsize=14, pad=10)

    # Add humidity statistics
    hum_min = min(humidities)
    hum_max = max(humidities)
    hum_avg = sum(humidities) / len(humidities)
    ax2.axhline(y=hum_avg, color="#4ecdc4", linestyle="--", alpha=0.7, label=f"Avg: {hum_avg:.1f}%")
    ax2.legend()

    # Format x-axis
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d %H:%M"))
    ax2.xaxis.set_major_locator(mdates.HourLocator(interval=6))
    plt.xticks(rotation=45)

    # Adjust layout
    plt.tight_layout()

    # Save to base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
    buffer.seek(0)
    chart_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()

    return chart_base64

def generate_pdf_html(weather_data, location_info, hours, chart_base64):
    """
    Generate HTML content for the PDF report.

    Args:
        weather_data: List of WeatherData objects
        location_info: String describing the location
        hours: Number of hours covered
        chart_base64: Base64 encoded chart image

    Returns:
        HTML string for PDF generation
    """
    # Calculate statistics
    temperatures = [record.temperature_2m for record in weather_data]
    humidities = [record.relative_humidity_2m for record in weather_data]

    temp_min = min(temperatures)
    temp_max = max(temperatures)
    temp_avg = sum(temperatures) / len(temperatures)

    hum_min = min(humidities)
    hum_max = max(humidities)
    hum_avg = sum(humidities) / len(humidities)

    # Get date range
    start_time = min(record.timestamp for record in weather_data)
    end_time = max(record.timestamp for record in weather_data)

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Weather Report</title>
    </head>
    <body>
        <div class="header">
            <div class="title">Weather Data Report</div>
            <p>Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>
        </div>

        <div class="metadata">
            <h3>Report Information</h3>
            <p><strong>Location:</strong> {location_info}</p>
            <p><strong>Time Range:</strong> {hours} hours</p>
            <p><strong>Data Period:</strong> {start_time.strftime("%Y-%m-%d %H:%M")} to {end_time.strftime("%Y-%m-%d %H:%M")}</p>
            <p><strong>Total Records:</strong> {len(weather_data)}</p>
        </div>

        <div class="chart-container">
            <div class="chart-title">Temperature and Humidity Trends</div>
            <img src="data:image/png;base64,{chart_base64}" style="max-width: 100%; height: auto;" />
        </div>

        <div class="summary">
            <div class="summary-title">Statistical Summary</div>
            <div style="display: flex; justify-content: space-between;">
                <div style="flex: 1; margin-right: 20px;">
                    <h4>Temperature (°C)</h4>
                    <p><strong>Minimum:</strong> {temp_min:.1f}°C</p>
                    <p><strong>Maximum:</strong> {temp_max:.1f}°C</p>
                    <p><strong>Average:</strong> {temp_avg:.1f}°C</p>
                    <p><strong>Range:</strong> {temp_max - temp_min:.1f}°C</p>
                </div>
                <div style="flex: 1;">
                    <h4>Relative Humidity (%)</h4>
                    <p><strong>Minimum:</strong> {hum_min:.1f}%</p>
                    <p><strong>Maximum:</strong> {hum_max:.1f}%</p>
                    <p><strong>Average:</strong> {hum_avg:.1f}%</p>
                    <p><strong>Range:</strong> {hum_max - hum_min:.1f}%</p>
                </div>
            </div>
        </div>

        <div style="margin-top: 30px; font-size: 12px; color: #666; text-align: center;">
            <p>Data source: Open-Meteo Archive API</p>
            <p>This report was generated automatically by the Weather Backend Service</p>
        </div>
    </body>
    </html>
    """

    return html_content

def generate_pdf_with_reportlab(weather_data, location_info, hours, chart_base64):
    """
    Fallback PDF generation using ReportLab when WeasyPrint is unavailable.
    Returns a file path to a generated PDF.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    import base64 as _b64
    import tempfile as _tmp
    from datetime import datetime as _dt

    width, height = A4
    temp = _tmp.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp.close()

    c = canvas.Canvas(temp.name, pagesize=A4)

    # Header
    c.setFillColor(colors.HexColor('#007acc'))
    c.setFont("Helvetica-Bold", 20)
    c.drawString(2*cm, height - 2*cm, "Weather Data Report")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, height - 2.7*cm, f"Generated on {_dt.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")

    # Metadata box
    y = height - 4*cm
    c.setStrokeColor(colors.HexColor('#007acc'))
    c.rect(1.8*cm, y - 3.2*cm, width - 3.6*cm, 3*cm, stroke=1, fill=0)
    c.setFont("Helvetica", 11)
    c.drawString(2.2*cm, y - 0.6*cm, f"Location: {location_info}")
    c.drawString(2.2*cm, y - 1.3*cm, f"Time Range: {hours} hours")
    start_time = min(record.timestamp for record in weather_data)
    end_time = max(record.timestamp for record in weather_data)
    c.drawString(2.2*cm, y - 2.0*cm, f"Data Period: {start_time.strftime('%Y-%m-%d %H:%M')} to {end_time.strftime('%Y-%m-%d %H:%M')}")
    c.drawString(2.2*cm, y - 2.7*cm, f"Total Records: {len(weather_data)}")

    # Chart image
    try:
        img_bytes = _b64.b64decode(chart_base64)
        img_temp = _tmp.NamedTemporaryFile(delete=False, suffix=".png")
        img_temp.write(img_bytes)
        img_temp.flush()
        img_temp.close()
        chart_w = width - 4*cm
        chart_h = 9*cm
        c.drawImage(img_temp.name, 2*cm, y - 3.2*cm - chart_h - 0.8*cm, width=chart_w, height=chart_h, preserveAspectRatio=True, anchor='sw')
        try:
            os.unlink(img_temp.name)
        except:
            pass
    except Exception:
        # If image fails, continue without chart
        pass

    # Summary statistics
    temperatures = [record.temperature_2m for record in weather_data]
    humidities = [record.relative_humidity_2m for record in weather_data]
    temp_min = min(temperatures)
    temp_max = max(temperatures)
    temp_avg = sum(temperatures) / len(temperatures)
    hum_min = min(humidities)
    hum_max = max(humidities)
    hum_avg = sum(humidities) / len(humidities)

    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(colors.HexColor('#007acc'))
    c.drawString(2*cm, 6*cm, "Statistical Summary")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 11)
    c.drawString(2*cm, 5.3*cm, f"Temperature (°C)  Min: {temp_min:.1f}   Max: {temp_max:.1f}   Avg: {temp_avg:.1f}")
    c.drawString(2*cm, 4.7*cm, f"Relative Humidity (%)  Min: {hum_min:.1f}   Max: {hum_max:.1f}   Avg: {hum_avg:.1f}")

    # Footer
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor('#666666'))
    c.drawCentredString(width/2, 1.8*cm, "Data source: Open-Meteo Archive API")
    c.drawCentredString(width/2, 1.4*cm, "This report was generated automatically by the Weather Backend Service")

    c.showPage()
    c.save()

    return temp.name
